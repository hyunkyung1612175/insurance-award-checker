import streamlit as st
import pdfplumber
import pytesseract
from pdf2image import convert_from_bytes
from PIL import Image
import pandas as pd
from io import BytesIO
from datetime import datetime
import re
from openpyxl.styles import Border, Side

# 허용된 사번 리스트
ALLOWED_IDS = ["1612175", "202301", "202302"]

# 인증 상태 초기화
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
if "user_id" not in st.session_state:
    st.session_state["user_id"] = ""
if "login_failed" not in st.session_state:
    st.session_state["login_failed"] = False

# 로그인 처리 함수
def authenticate(user_id):
    if user_id in ALLOWED_IDS:
        st.session_state["authenticated"] = True
        st.session_state["user_id"] = user_id
        st.session_state["login_failed"] = False
    else:
        st.session_state["authenticated"] = False
        st.session_state["user_id"] = ""
        st.session_state["login_failed"] = True

# 로그인 화면
if not st.session_state["authenticated"]:
    st.markdown("""
        <style>
        .login-box {
            background-color: #0056b3;
            padding: 40px;
            border-radius: 10px;
            box-shadow: 2px 2px 10px rgba(0,0,0,0.2);
            max-width: 500px;
            margin: auto;
            text-align: center;
            color: white;
        }
        .login-box h1 {
            font-size: 48px;
            margin-bottom: 0;
        }
        .login-box h3 {
            font-size: 24px;
            margin-top: 0;
        }
        </style>
        <div class="login-box">
            <h1>생보관리팀</h1>
            <h3>보험사 시상 확인</h3>
        </div>
    """, unsafe_allow_html=True)

    user_id_input = st.text_input("사번:", value="")
    if st.button("로그인"):
        authenticate(user_id_input)

    if st.session_state["login_failed"]:
        st.markdown("""
            <div style="background-color:#ffe6e6; padding:20px; border-radius:10px; border:1px solid #ff4d4d;">
                <h4 style="color:#cc0000;">❌ 허용되지 않은 사용자입니다.</h4>
                <p>관리자에게 문의해주세요.</p>
            </div>
        """, unsafe_allow_html=True)

    st.stop()

# 기능 화면 제목
st.markdown("""
    <div style="text-align:center; margin-bottom:30px;">
        <h1 style="font-size:48px; color:#2c3e50; margin-bottom:0;">생보관리팀</h1>
        <h3 style="font-size:24px; color:#555555; margin-top:0;">보험사 시상 확인</h3>
    </div>
""", unsafe_allow_html=True)

# 기능 선택 메뉴
menu = st.radio("기능 선택", ["📥 PDF 업로드", "📊 엑셀 변환", "✏️ 내용 수정"])

# 텍스트 추출 함수
def extract_text_from_pdf(file):
    text = ""
    try:
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except:
        images = convert_from_bytes(file.read())
        for img in images:
            text += pytesseract.image_to_string(img, lang='kor+eng') + "\n"

    # "추가시책" 이후 내용은 잘라내기
    cutoff = text.find("추가시책")
    if cutoff != -1:
        text = text[:cutoff]

    return text

SPECIAL_MAP = {
    "간편심사": "(간편)",
    "일반심사": "(일반)",
    "달러": "(달러)",
    "원화": "(원화)"
}

def parse_text_to_tables(text):
    lines = text.splitlines()
    data_bonsa, data_jisa, data_fp = [], [], []

    SPECIAL_MAP = {
        "간편심사": "(간편)",
        "일반심사": "(일반)",
        "달러": "(달러)",
        "원화": "(원화)"
    }

    for i, line in enumerate(lines):
        if re.search(r"\d+%|-", line):  # 퍼센트나 "-" 있는 줄만 처리
            parts = line.split()

            # 상품명 추출 (%나 "-" 나오기 전까지)
            name_tokens = []
            for token in parts:
                if "%" in token or token == "-":
                    break
                name_tokens.append(token)
            name = " ".join(name_tokens)

            # 규칙 1: 상품명이 3자 이하라면 윗줄 끌어오기
            if len(name) <= 3 and i > 0:
                prev_line = lines[i-1].strip()
                if not re.search(r"\d+%", prev_line):  # 윗줄에 % 없음
                    name = prev_line + " " + name

            # 규칙 2: 간편심사 / 일반심사 / 원화 / 달러 처리
            if name in SPECIAL_MAP and i > 0:
                prev_line = lines[i-1].strip()
                if re.search(r"\d+%", prev_line):  # 윗줄에도 %가 있는 경우
                    # 윗줄 상품명 추출
                    prev_tokens = []
                    for token in prev_line.split():
                        if "%" in token or token == "-":
                            break
                        prev_tokens.append(token)
                    prev_name = " ".join(prev_tokens)

                    # 괄호나 심사 토큰 제거
                    prev_name = re.sub(r"\(.*?\)", "", prev_name).strip()
                    prev_name = prev_name.replace("일반심사", "").replace("간편심사", "").strip()

                    # 특별 케이스: 원화/달러
                    if name == "원화":
                        if "달러" in prev_name:
                            prev_name = prev_name.split("달러")[0].strip()
                        name = prev_name + SPECIAL_MAP[name]
                    elif name == "달러":
                        if "원화" in prev_name:
                            prev_name = prev_name.split("원화")[0].strip()
                        name = prev_name + SPECIAL_MAP[name]
                    else:
                        name = prev_name + SPECIAL_MAP[name]

            # 괄호 처리: 다음 줄에 () 있으면 상품명에 추가
            if i+1 < len(lines):
                next_line = lines[i+1].strip()
                if next_line.startswith("(") and next_line.endswith(")"):
                    name = name + next_line

            # 남은 부분
            remainder = parts[len(name_tokens):]
            values = [p for p in remainder if "%" in p or p == "-"]

            if len(values) >= 6:
                bonsa_익월, bonsa_13 = values[0], values[1]
                jisa_익월, jisa_13 = values[2], values[3]
                fp_익월, fp_13 = values[4], values[5]

                data_bonsa.append([name, bonsa_익월, bonsa_13])
                data_jisa.append([name, jisa_익월, jisa_13])
                data_fp.append([name, fp_익월, fp_13])

    return data_bonsa, data_jisa, data_fp

# PDF 업로드 및 변환
if menu == "📥 PDF 업로드":
    st.markdown("### PDF 파일 업로드 및 변환")
    uploaded_file = st.file_uploader("PDF 파일을 선택하세요", type=["pdf"])

    if uploaded_file:
        raw_text = extract_text_from_pdf(uploaded_file)
        st.subheader("📑 추출된 텍스트")
        st.text_area("PDF 내용", raw_text, height=300)

        bonsa_data, jisa_data, fp_data = parse_text_to_tables(raw_text)

        st.subheader("📊 [본사] 정리된 표")
        df_bonsa = pd.DataFrame(bonsa_data, columns=["상품명", "익월", "13회차"])
        st.dataframe(df_bonsa)

        st.subheader("📊 [지사] 정리된 표")
        df_jisa = pd.DataFrame(jisa_data, columns=["상품명", "익월", "13회차"])
        st.dataframe(df_jisa)

        st.subheader("📊 [FP] 정리된 표")
        df_fp = pd.DataFrame(fp_data, columns=["상품명", "익월", "13회차"])
        st.dataframe(df_fp)

        # 엑셀 저장 (테두리 + A열 자동폭 적용)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_bonsa.to_excel(writer, sheet_name="본사", index=False)
            df_jisa.to_excel(writer, sheet_name="지사", index=False)
            df_fp.to_excel(writer, sheet_name="FP", index=False)

            workbook = writer.book
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for sheet_name in ["본사", "지사", "FP"]:
                sheet = workbook[sheet_name]

                # 테두리 적용
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                           min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.border = thin_border

                # A열 폭 자동 조정
                max_length = 0
                for cell in sheet["A"]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions["A"].width = max_length + 2

        st.download_button(
            label="📥 엑셀 파일 다운로드",
            data=output.getvalue(),
            file_name="정리된_시상표.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# 엑셀 변환 (추가 기능 자리)
elif menu == "📊 엑셀 변환":
    st.markdown("### 엑셀 변환 기능")
    st.button("엑셀로 변환 시작")

# 내용 수정
elif menu == "✏️ 내용 수정":
    st.markdown("### 내용 수정")
    new_content = st.text_area("수정할 내용을 입력하세요")
    if st.button("수정 저장"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        history = st.session_state.get("history", [])
        history.append({
            "사번": st.session_state["user_id"],
            "날짜": timestamp,
            "내용": new_content
        })
        st.session_state["history"] = history
        st.success("수정 내용이 저장되었습니다.")
